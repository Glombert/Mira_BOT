"""
tools/excel_tools.py — инструменты для работы с Excel-файлами.

Простая версия: чтение и запись данных без форматирования.
Все пути изолированы внутри workspace/{user_id}/ через _resolve_path().

Функции:
    excel_read()  — читает лист, возвращает заголовки + строки
    excel_write() — создаёт .xlsx с заголовками и данными
"""

import os
from tools.file_tools import _resolve_path, _save_undo

MAX_ROWS = 200  # максимум строк чтобы не засорять контекст модели


def excel_read(user_id: str, relative_path: str,
               sheet_name: str | None = None,
               max_rows: int = MAX_ROWS) -> dict:
    """
    Читает Excel-файл из workspace/{user_id}/.

    Аргументы:
        user_id       — идентификатор пользователя
        relative_path — путь к .xlsx относительно workspace
        sheet_name    — имя листа; если не указан — берётся первый
        max_rows      — максимум строк данных (без заголовка)

    Возвращает:
        {
          "ok": True,
          "sheet":   "Лист1",
          "sheets":  ["Лист1", "Лист2"],
          "headers": ["Имя", "Возраст", ...],
          "rows":    [["Иван", 30], ["Мария", 25], ...],
          "truncated": False   # True если строк больше max_rows
        }
    """
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl не установлен. Запусти: pip install openpyxl"}

    try:
        target = _resolve_path(user_id, relative_path)
    except PermissionError as e:
        return {"ok": False, "error": str(e)}

    if not os.path.exists(target):
        return {"ok": False, "error": f"Файл '{relative_path}' не найден."}
    if not os.path.isfile(target):
        return {"ok": False, "error": f"'{relative_path}' — это папка, не файл."}
    if not target.lower().endswith((".xlsx", ".xlsm", ".xltx")):
        return {"ok": False, "error": "Поддерживаются только .xlsx файлы."}

    try:
        wb = openpyxl.load_workbook(target, read_only=True, data_only=True)
        sheet_names = wb.sheetnames

        if sheet_name:
            if sheet_name not in sheet_names:
                wb.close()
                return {
                    "ok": False,
                    "error": f"Лист '{sheet_name}' не найден. Доступные: {sheet_names}"
                }
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        rows_iter = ws.iter_rows(values_only=True)

        # Первая строка — заголовки
        try:
            raw_headers = next(rows_iter)
        except StopIteration:
            wb.close()
            return {
                "ok": True, "sheet": sheet_name, "sheets": sheet_names,
                "headers": [], "rows": [], "truncated": False,
            }

        headers = [str(h) if h is not None else "" for h in raw_headers]

        # Читаем строки данных до max_rows
        rows = []
        truncated = False
        for row in rows_iter:
            if len(rows) >= max_rows:
                truncated = True
                break
            rows.append([
                val if val is not None else ""
                for val in row
            ])

        wb.close()
        return {
            "ok":        True,
            "sheet":     sheet_name,
            "sheets":    sheet_names,
            "headers":   headers,
            "rows":      rows,
            "truncated": truncated,
        }

    except Exception as e:
        return {"ok": False, "error": f"Ошибка чтения файла: {e}"}


def excel_write(user_id: str, relative_path: str,
                headers: list, rows: list,
                sheet_name: str = "Sheet1",
                overwrite: bool = False) -> dict:
    """
    Создаёт Excel-файл в workspace/{user_id}/.

    Аргументы:
        user_id       — идентификатор пользователя
        relative_path — путь к .xlsx относительно workspace
        headers       — список заголовков: ["Имя", "Возраст", ...]
        rows          — список строк: [["Иван", 30], ["Мария", 25], ...]
        sheet_name    — имя листа (по умолчанию "Sheet1")
        overwrite     — разрешить перезапись существующего файла

    Возвращает:
        {"ok": True, "path": "output/report.xlsx", "rows_written": 5}
    """
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl не установлен. Запусти: pip install openpyxl"}

    try:
        target = _resolve_path(user_id, relative_path)
    except PermissionError as e:
        return {"ok": False, "error": str(e)}

    if os.path.exists(target) and not overwrite:
        return {
            "ok": False,
            "error": (
                f"Файл '{relative_path}' уже существует. "
                "Передай overwrite=true чтобы перезаписать."
            )
        }

    # Бэкап перед перезаписью
    if os.path.exists(target) and overwrite:
        _save_undo(user_id, relative_path, target)

    os.makedirs(os.path.dirname(target), exist_ok=True)

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Заголовки жирным
        from openpyxl.styles import Font
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Данные
        for row in rows:
            ws.append(list(row))

        wb.save(target)

        return {
            "ok":          True,
            "path":        relative_path,
            "rows_written": len(rows),
        }

    except Exception as e:
        return {"ok": False, "error": f"Ошибка создания файла: {e}"}
